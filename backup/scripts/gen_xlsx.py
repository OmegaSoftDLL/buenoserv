#!/usr/bin/env python3
"""Gerador de planilhas Excel (.xlsx) — openpyxl"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils import get_column_letter
import sys, json, os

STYLE_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
STYLE_HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
STYLE_TITLE_FONT = Font(bold=True, size=14, color="1A237E", name='Calibri')
STYLE_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
STYLE_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def nova_planilha():
    return openpyxl.Workbook()

def salvar(wb, path):
    wb.save(path)
    print(f"XLSX|OK|{path}")

def cabecalho(ws, dados, row=1):
    for col, h in enumerate(dados, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = STYLE_HEADER_FONT
        c.fill = STYLE_HEADER_FILL
        c.alignment = STYLE_CENTER
        c.border = STYLE_BORDER

def linha(ws, dados, row):
    for col, v in enumerate(dados, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.border = STYLE_BORDER
        c.alignment = STYLE_CENTER

def auto_largura(ws, cab):
    for i, h in enumerate(cab, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(str(h)) + 2, 15)

def gerar_tabela(nome, sheet_name, cabecalhos, dados):
    wb = nova_planilha()
    ws = wb.active
    ws.title = sheet_name
    cabecalho(ws, cabecalhos)
    for i, row in enumerate(dados, 2):
        linha(ws, row, i)
    auto_largura(ws, cabecalhos)
    salvar(wb, nome)

def gerar_curva_s(nome, meses, planejado, realizado, titulo="Curva S"):
    wb = nova_planilha()
    ws = wb.active
    ws.title = "Curva S"
    cabecalho(ws, ["Mês", "Planejado (R$)", "Realizado (R$)", "% Físico"])
    for i, m in enumerate(meses):
        perc = round(realizado[i] / planejado[-1] * 100, 1) if planejado[-1] else 0
        ws.append([m, planejado[i], realizado[i], perc])
    chart = LineChart()
    chart.title = titulo
    chart.y_axis.title = "Valor Acumulado (R$)"
    chart.x_axis.title = "Mês"
    chart.style = 10
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(meses)+1)
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=len(meses)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 20
    ws.add_chart(chart, "E2")
    salvar(wb, nome)

def gerar_cronograma(nome, tarefas):
    cab = ["ID","EAP","Tarefa","Duração (dias)","Início","Término",
           "Predecessora","Recurso","Custo Material","Custo MO",
           "% Planejado","% Realizado","Status"]
    wb = nova_planilha()
    ws = wb.active
    ws.title = "Cronograma"
    cabecalho(ws, cab)
    for i, t in enumerate(tarefas, 2):
        linha(ws, t, i)
    ws.column_dimensions['C'].width = 45
    auto_largura(ws, cab)
    salvar(wb, nome)

def gerar_dre_xlsx(nome, dre_data, mes, ano):
    wb = nova_planilha()
    ws = wb.active
    ws.title = f"DRE {mes:02d}/{ano}"
    
    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value=f"BUENOSERV — DRE {mes:02d}/{ano}").font = STYLE_TITLE_FONT
    
    itens = [
        ("1. Receita Bruta de Serviços", dre_data.get("receita_bruta", 0)),
        ("2. Deduções", dre_data.get("deducoes", 0)),
        ("", None),
        ("3. Receita Líquida", dre_data.get("receita_liquida", 0)),
        ("4. Custos dos Serviços", dre_data.get("custos_servicos", 0)),
        ("", None),
        ("5. Lucro Bruto", dre_data.get("lucro_bruto", 0)),
        ("6. Despesas Administrativas", dre_data.get("despesas_administrativas", 0)),
        ("7. Despesas Comerciais", dre_data.get("despesas_comerciais", 0)),
        ("8. Despesas Tributárias", dre_data.get("despesas_tributarias", 0)),
        ("", None),
        ("9. Lucro Operacional", dre_data.get("lucro_operacional", 0)),
        ("10. Resultado Financeiro", dre_data.get("resultado_financeiro", 0)),
        ("", None),
        ("11. Lucro Líquido", dre_data.get("lucro_liquido", 0)),
    ]
    
    ws.cell(row=2, column=1, value="Conta").font = Font(bold=True, size=10)
    ws.cell(row=2, column=2, value="Valor (R$)").font = Font(bold=True, size=10)
    ws.cell(row=2, column=1).fill = STYLE_HEADER_FILL
    ws.cell(row=2, column=2).fill = STYLE_HEADER_FILL
    ws.cell(row=2, column=1).font = STYLE_HEADER_FONT
    ws.cell(row=2, column=2).font = STYLE_HEADER_FONT
    
    for i, (label, valor) in enumerate(itens, 3):
        if label:
            ws.cell(row=i, column=1, value=label)
            c = ws.cell(row=i, column=2, value=valor)
            c.number_format = '#,##0.00'
            if "Lucro" in label and "Operacional" not in label:
                ws.cell(row=i, column=1).font = Font(bold=True, size=11)
                c.font = Font(bold=True, size=11)
            if "Líquido" in label:
                ws.cell(row=i, column=1).font = Font(bold=True, size=12, color="1A237E")
                c.font = Font(bold=True, size=12, color="1A237E")
    
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 18
    salvar(wb, nome)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: gen_xlsx.py <comando> <args_json>")
        print("Comandos: tabela, curva_s, cronograma")
        sys.exit(1)
    cmd = sys.argv[1]
    args = json.loads(sys.argv[2]) if len(sys.argv) > 2 else {}
    if cmd == "tabela":
        gerar_tabela(args["nome"], args.get("sheet","Dados"), args["cabecalhos"], args["dados"])
    elif cmd == "curva_s":
        gerar_curva_s(args["nome"], args["meses"], args["planejado"], args["realizado"], args.get("titulo","Curva S"))
    elif cmd == "cronograma":
        gerar_cronograma(args["nome"], args["tarefas"])
    else:
        print(f"Comando desconhecido: {cmd}")
