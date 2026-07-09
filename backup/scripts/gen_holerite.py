#!/usr/bin/env python3
"""Gerador de holerite XLSX"""
import os, sys, json, datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl necessário")
    sys.exit(1)

STATE_FILE = os.path.expanduser("~/.config/opencode/state/agent_state.json")

def gerar_holerite(funcionario, salario_bruto, descontos=None, saida=None):
    descontos = descontos or {"INSS": salario_bruto * 0.11, "IRRF": salario_bruto * 0.075}
    saida = saida or os.path.expanduser(f"~/Desktop/Holerite_{funcionario.replace(' ','_')}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Holerite {datetime.date.today().month:02d}/{datetime.date.today().year}"

    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value=f"BUENOSERV — Holerite").font = Font(bold=True, size=14, color="1A237E")
    ws.cell(row=2, column=1, value=f"Funcionário: {funcionario}").font = Font(bold=True, size=11)
    ws.cell(row=3, column=1, value=f"Período: {datetime.date.today():%B/%Y}").font = Font(size=10)

    cab = ["Descrição", "Referência", "Valor (R$)", ""]
    for col, h in enumerate(cab, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")

    rows = [
        ("Salário Bruto", "", salario_bruto),
        *[(f"Desconto: {k}", "", v) for k, v in descontos.items()],
        ("", "", ""),
        ("Salário Líquido", "", salario_bruto - sum(descontos.values())),
    ]
    for i, (desc, ref, valor) in enumerate(rows, 6):
        ws.cell(row=i, column=1, value=desc)
        ws.cell(row=i, column=2, value=ref)
        c = ws.cell(row=i, column=3, value=valor)
        c.number_format = '#,##0.00'
        if "Líquido" in desc:
            ws.cell(row=i, column=1).font = Font(bold=True, size=11, color="1A237E")
            c.font = Font(bold=True, size=11, color="1A237E")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['C'].width = 15
    wb.save(saida)
    print(f"✅ Holerite gerado: {saida}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: gen_holerite.py <funcionario> <salario_bruto>")
        sys.exit(1)
    gerar_holerite(sys.argv[1], float(sys.argv[2]))
